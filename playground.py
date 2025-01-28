from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule, FormulaRule

# Open the .xlsm workbook
file_path = "example.xlsm"
workbook = load_workbook(file_path, keep_vba=True)
sheet = workbook.active  # Select the active worksheet

# Define a fill for conditional formatting
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# Add conditional formatting for cells greater than 50
sheet.conditional_formatting.add(
    "A1:A10",  # Range to apply the rule
    CellIsRule(operator="greaterThan", formula=["50"], stopIfTrue=True, fill=green_fill)
)

# Add conditional formatting for even numbers using a formula
sheet.conditional_formatting.add(
    "B1:B10",  # Range to apply the rule
    FormulaRule(formula=["MOD(ROW(), 2) = 0"], stopIfTrue=True, fill=red_fill)
)

# Save the workbook
workbook.save("example_with_conditional_formatting.xlsm")
