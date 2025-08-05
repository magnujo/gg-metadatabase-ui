from openpyxl.formatting.rule import CellIsRule, FormulaRule
from PIL import ImageFont, ImageDraw, Image
import math
import os
from utils.misc import calculate_text_box_height, calculate_text_width
from openpyxl.comments import Comment
import pandas as pd
import psycopg2
from constants.misc_constants import SCRIPT_GENERATED_COLUMNS, ADMIN_EMAIL
from constants.db_names.name_maps import db_to_sheet_rename_map, sheet_to_db_rename_map
from constants.db_names.names import data
from constants.db_connections import ENGINE_READ_ONLY, DATABASE_CONFIG_READ_ONLY, PSY_CONN
from utils import misc
from openpyxl.utils import get_column_letter
from pathlib import Path

def find_project_root():
    path = Path(__file__).resolve()
    while path != path.root:
        if (path / 'very_rootsy_file.txt').exists():
            return path
        path = path.parent
    return None  # Project root not found

project_root = find_project_root()


from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from openpyxl.worksheet.datavalidation import DataValidation


"""
Function gets called when clicking on:
    http://dandyweb01fl.unicph.domain:5100/download/Field%20Sampling%20Meta%20data%20reporting%20template.xlsx
"""
def generate(table_name, schema_name, conn): 
    
    mandatory_colour = PatternFill(start_color='8ED973', end_color='8ED973', fill_type='solid')
    non_mandatory_colour = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    feature_dependent_colour = PatternFill(start_color='D9EFCD', end_color='D9EFCD', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(top=Side('thin'), bottom=Side('thin'), left=Side('thin'), right=Side('thin'))
    warning_fill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid')

    file_name = 'Field Sampling Meta data reporting template'
    extension = '.xlsm'
    save_path = os.path.join(project_root, 'auto_generated_upload_sheet_templates', file_name + extension)


    table_name = data.field_sample()
    schema_name = data()
    max_rows = 1048576

    if table_name == data.field_sample():
        col_names = data.field_sample

    # Query a good example row
    query = f'''
    SELECT * FROM "{schema_name}"."{table_name}" where "{col_names.field_sample_id()}" = 'min22a_3';
    '''

    df = pd.read_sql(query, conn)
    
    df = df.drop(columns=SCRIPT_GENERATED_COLUMNS, errors='ignore')

    context_types = pd.read_sql(f'select * from "{schema_name}"."{data.field_sample_context_types()}"', con=ENGINE_READ_ONLY)[data.field_sample_context_types.name()]
    environment_types = pd.read_sql(f'select * from "{schema_name}"."{data.field_sample_environment_types()}"', con=ENGINE_READ_ONLY)[data.field_sample_environment_types.name()]
    material_types = pd.read_sql(f'select * from "{schema_name}"."{data.field_sample_material_type()}"', con=ENGINE_READ_ONLY)[data.field_sample_material_type.name()]
    sample_types = pd.read_sql(f'select * from "{schema_name}"."{data.field_sample_types()}"', con=ENGINE_READ_ONLY)[data.field_sample_types.name()]
    country_ocean = pd.read_sql(f'select * from "{schema_name}"."{data.country_ocean()}"', con=ENGINE_READ_ONLY)[data.country_ocean.name()]
    field_sample_types_gm = pd.read_sql(f'select * from "{schema_name}"."{data.field_sample_types_gm()}"', con=ENGINE_READ_ONLY)[data.field_sample_types_gm.name()]

    enum_sheet = pd.DataFrame({
        data.field_sample.sample_context(template=True): context_types,
        data.field_sample.sample_environment(template=True): environment_types,
        data.field_sample.sample_environment_secondary(template=True): environment_types,
        data.field_sample.sample_material(template=True): material_types,
        data.field_sample.sample_type(template=True): sample_types,
        data.field_sample.sample_type_in_storage_at_gm(template=True): field_sample_types_gm,
        data.field_sample.country_ocean(template=True): country_ocean
    })

    for col in df.select_dtypes(include='bool').columns:
        df[col] = df[col].map({True: 'yes', False: 'no'})
    
    # Order that columns will appear in sheet 
    new_order = [
        col_names.field_sample_id(template=True),
        col_names.parent_id(template=True),
        col_names.field_label(template=True),
        col_names.master_id_parent_sample_id(template=True),
        col_names.running_project_title(template=True),
        col_names.permit_for_dna_analysis(template=True),
        col_names.country_ocean(template=True),
        col_names.site_name(template=True),
        col_names.geographical_location_names(template=True),
        col_names.latitude(template=True),
        col_names.longitude(template=True),
        col_names.elevation(template=True),
        col_names.water_depth(template=True),
        col_names.sample_context(template=True),
        col_names.sample_type(template=True),
        col_names.sample_type_in_storage_at_gm(template=True),
        col_names.sample_material(template=True),
        col_names.sample_environment(template=True),
        col_names.sample_environment_secondary(template=True),
        col_names.age_estimate___from(template=True),
        col_names.age_estimate___to(template=True),
        col_names.sampling_depth(template=True),
        col_names.sampling_interval___from(template=True),
        col_names.sampling_interval___to(template=True),
        col_names.sample_date(template=True),
        col_names.pi(template=True),
        col_names.sample_provider_name(template=True),
        col_names.sample_provider_contact(template=True),
        col_names.owned_by_aegis(template=True),
        col_names.sample_storage_setting(template=True),
        col_names.sample_storage_location(template=True),
        col_names.sample_storage_address(template=True),
        col_names.comments(template=True),
        col_names.link_to_other_relevant_information(template=True),
        col_names.link_to_images(template=True),
        col_names.miscellaneous_environmental_measurement_or_observation(template=True),
        col_names.miscellaneous_sample_measurement_or_observation(template=True),
        col_names.alias(template=True),
        col_names.cultural_affiliation(template=True),
        col_names.museum_institution(template=True),
        col_names.site_grid_elev(template=True),
        col_names.site_grid_latitude(template=True),
        col_names.site_grid_longitude(template=True),
        col_names.other_relevant_information(template=True)

    ]

    renamer = db_to_sheet_rename_map(schema_name=schema_name, table_name=table_name)
    df = df.rename(columns=renamer, errors='raise')
    
    df = df[new_order]
    

    new_rows = [
        (1, ""),
        (0, "Colour legend:"),
        (1, " = Bad value. For example if a categorical value is not in the Allowed categorical values sheet or if a latitude value is not between -90 and 90"),
        (1, " = Mandatory column"),
        (1, " = Non-mandatory column"),
        (1, " = Mandatory column depending on environment, type or other features"),
        (0, '''
    IMPORTANT: If you get a warning about macros being blocked please do the necessary steps described here to unblock, otherwise you might end up with wrong or badly formatted data 
    https://support.microsoft.com/en-us/topic/a-potentially-dangerous-macro-has-been-blocked-0952faa0-37e7-4316-b61d-5b5ed6024216
            '''),
        (0, ""),
        (0, "Delete this and all rows above before uploading/sending - except the header (row 1) ofcourse!")
]
    
    temp_row = df.loc[0]
    df.loc[0] = ["EXAMPLE ROW:"] + ([None] * (len(df.columns)-1)) 
    df.loc[1] = temp_row
    # df.loc[0] = ["EXAMPLE ROW:"] + ([None] * (len(df.columns)-1)) 
    for i, row in enumerate(new_rows):
        k, v = row
        new_row = [None] * len(df.columns)
        new_row[k] = v
        df.loc[(len(df))] = new_row
        
    # df_list = df.to_dict('records')

    # # Insert rows at specific positions
   
    # df_list.insert(0, new_row_0)
    
    # for i, row in enumerate(new_rows):
    #     df_list.insert(i+2, row)

    # # Convert back to DataFrame
    # df = pd.DataFrame(df_list)

    # Sort the index to maintain order
    df = df.sort_index().reset_index(drop=True)

    # Handle timezone-aware datetime columns
    # for col in df.columns:
    #     if pd.api.types.is_datetime64tz_dtype(df[col]):
    #         df[col] = df[col].dt.tz_localize(None)

    for col in df.columns:
        if isinstance(df[col].dtype, pd.DatetimeTZDtype):
            df[col] = df[col].dt.tz_localize(None)
            
    # Query to fetch column constraints
    constraints_query = f"""
    SELECT column_name, is_nullable
    FROM information_schema.columns
    WHERE table_name = '{table_name}' and table_schema = '{schema_name}'
    """
    constraints_df = pd.read_sql(constraints_query, conn)

    # Drop auto-generated columns
    df = df.drop(columns=SCRIPT_GENERATED_COLUMNS, errors='ignore')

    # Rename columns based on the renaming map
    renamer = db_to_sheet_rename_map(schema_name=schema_name, table_name=table_name)
    # df_translated = df.rename(columns=renamer, errors='raise')

    # New column order

    df_translated = df[new_order]

    input_file_path = os.path.join(project_root, 'auto_generated_upload_sheet_templates', 'blankfile.xlsm')
        # Load the existing .xlsm file
    workbook = load_workbook(input_file_path, keep_vba=True)

    data_sheet = 'Insert Data Here'

    data_map = {'Insert Data Here': df_translated, 'Allowed categorical values': enum_sheet}
    
    # Access the active sheet (or you can specify a sheet by name)
    for sheet, df in data_map.items():
    
        sheet = workbook[sheet]
    
        # Write DataFrame headers to the first row
        for col_num, header in enumerate(df.columns, start=1):
            sheet.cell(row=1, column=col_num, value=header)

        # Write DataFrame values to the sheet
        for row_num, row in enumerate(df.values, start=2):  # Start writing from the second row
            for col_num, value in enumerate(row, start=1):
                sheet.cell(row=row_num, column=col_num, value=value)
                

    
    # Create a Pandas Excel writer using openpyxl as the engine
    
    enum_sheet_name = 'Allowed categorical values'
    data_sheet_name = 'Insert Data Here'
    
    # Write the DataFrame to Excel
    
    # Create a new sheet for the guide
    guide_sheet = workbook['Read this first!']
    
    # Add instructions to the guide sheet
    instructions = [
        "To limit data insertion errors, please read the instructions below before you begin to insert data:",
        f"1. Ask yourself if you are using the correct template. The purpose of this template is to provide meta data about field samples i.e. unprocessed samples that were collected from a field sample environment (see list of sample environments in the '{enum_sheet_name}' sheet). It is also possible to report meta data about sub field samples using this template, as long as they are not archive- or robot sub samples (in LV tubes). These are special samples that have their own templates. Contact jtstenderup@sund.ku.dk if you need the templates for those.",
        "2. When inserting data in a column, hover over the column name to see its definition. Read the column definition thoroughly before inserting the data, to make sure you are inserting it in the correct column",
        f"3. There are some columns where you will see a drop-down list when you click on a cell. This means that you can only insert values from that list. You can also see the lists in the '{enum_sheet_name}' sheet. If you wish to include a value to one of these lists, contact {ADMIN_EMAIL} (it wont help if you just add it to the template yourself)", 
        "4. Before uploading/sending the file, ensure all entries are accurate and complete. It's better to leave a cell empty than to insert a wrong value!",
        "5. Delete any empty rows and columns",
        "6. Delete the yellow row and all the rows above it, except the row with the header.",
        f"Feel free to contact {ADMIN_EMAIL} if you have any questions or feedback to this template"
    ]

    for idx, instruction in enumerate(instructions, start=1):
        guide_sheet[f'A{idx}'] = instruction

    
    # Optionally, adjust column width for better readability
    for col in guide_sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        guide_sheet.column_dimensions[column].width = adjusted_width
    
    
    worksheet = workbook[data_sheet_name]
    # bold_format = workbook.add_format({'bold': True})
    
    # Apply the dropdown validation and conditional formatting
    
    num_of_rows = 1000  # Number of rows to apply dropdown and conditional formatting to
    start_row = 2  # The row where the data begins (ie. not including column header)

    for i, col in enumerate(df_translated.columns):
        column_letter_data = get_column_letter(i + 1)
        if col == data.field_sample.latitude(template=True):
            rule = FormulaRule(formula=[f"=OR({column_letter_data}{start_row}< -90, {column_letter_data}{start_row} > 90)"], stopIfTrue=True, fill=warning_fill)
            worksheet.conditional_formatting.add(f"{column_letter_data}{start_row}:{column_letter_data}{num_of_rows}", rule)  # Adjust the range accordingly
        
        if col == data.field_sample.longitude(template=True):
            rule = FormulaRule(formula=[f"=OR({column_letter_data}{start_row}< -180, {column_letter_data}{start_row} > 180)"], stopIfTrue=True, fill=warning_fill)
            worksheet.conditional_formatting.add(f"{column_letter_data}{start_row}:{column_letter_data}{num_of_rows}", rule)  # Adjust the range accordingly

        
        if col in enum_sheet.columns:
            dropdown_values = list(enum_sheet[col].dropna())
            formula = f'"{",".join(dropdown_values)}"'

            
            column_index = enum_sheet.columns.get_loc(col)
            enums_length = len(dropdown_values)
            col_letter_enum = get_column_letter(column_index + 1)
            formula = f"'Allowed categorical values'!${col_letter_enum}${start_row}:${col_letter_enum}${enums_length+1}"
            dropdown = DataValidation(
                type="list",
                formula1=formula,
                allow_blank=False,
                showErrorMessage=True,  # Enable error message
                errorTitle="Invalid Value",
                error="The value you entered is not in the allowed list. Please select a valid value from the dropdown."
            )
            worksheet.add_data_validation(dropdown)
            for row in range(start_row, num_of_rows):  # Excel rows start at 1, header is row 1
                cell = f'{column_letter_data}{row}'  # 'B' is the second column (Choice column)
                dropdown.add(worksheet[cell])

            # Define the conditional formatting rule
            formula = f'=AND({column_letter_data}{start_row}<>"", COUNTIF({formula}, {column_letter_data}2)=0)'

            # Apply the formula rule to a range of cells (adjust range as needed)
            rule = FormulaRule(formula=[formula], stopIfTrue=True, fill=warning_fill)

            
            worksheet.conditional_formatting.add(f"{column_letter_data}{start_row}:{column_letter_data}{num_of_rows}", rule)  # Adjust the range accordingly
        
    # Define header formats



    # Map renamed columns back to the original names for constraint checking
    reverse_renamer = {v: k for k, v in renamer.items()}

    special_non_mandatory_columns = [
        col_names.sampling_depth(),
        col_names.sampling_interval___from(),
        col_names.sampling_interval___to(),
        col_names.sample_environment_secondary(),
        col_names.parent_id()
    ]

    #  Columns that are not mandatory in DB but actually should be filled
    should_be_mandatory = [
        col_names.sample_date(),
    ]

    comments = misc.get_comments(DATABASE_CONFIG_READ_ONLY['dbname'], data(), data.field_sample(), psy_conn=PSY_CONN)
    
    for col_num, col_name in enumerate(df_translated.columns):
        # Get the original column name
        original_col_name = reverse_renamer.get(col_name, col_name)

        # Check the NOT NULL constraint
        is_nullable = constraints_df[constraints_df['column_name'] == original_col_name]['is_nullable'].values[0]
        
        # Apply color based on constraints and special non-mandatory status
        cell = worksheet.cell(row=1, column=col_num + 1)
        if original_col_name in special_non_mandatory_columns:
            cell.fill = feature_dependent_colour
        elif is_nullable == 'NO' or original_col_name in should_be_mandatory:
            cell.fill = mandatory_colour
        else:
            cell.fill = non_mandatory_colour
        
        comment = str(comments[comments['Column Name'] == original_col_name]['Comment'].iloc[0])
        
    
        char_width = 7
        text_width = len(comment) * char_width 
        line_width = 40 * char_width
        n_lines = math.ceil(text_width / line_width)
        char_height = 22
        
        height = (char_height * n_lines) + char_height
        
        new_lines = comment.count("\n")
        height = height + (char_height * new_lines) 
    
        cell.comment = Comment(comment, ADMIN_EMAIL, width=line_width, height=height)
        
        # Align the text in the cell
        cell.alignment = center_alignment
    
        macro_warning_cell = worksheet.cell(row=10, column=col_num + 1)
        example_cell = worksheet.cell(row=12, column=col_num + 1)
        
        example_cell.fill = yellow_fill
        example_cell.font = Font(bold=True)
        macro_warning_cell.fill = warning_fill
        macro_warning_cell.font = Font(bold=True)

    worksheet.cell(row=6, column=1).fill = warning_fill 
    worksheet.cell(row=6, column=1).border = border 
    worksheet.cell(row=7, column=1).fill = mandatory_colour
    worksheet.cell(row=7, column=1).border = border 
    worksheet.cell(row=8, column=1).fill = non_mandatory_colour 
    worksheet.cell(row=8, column=1).border = border 
    worksheet.cell(row=9, column=1).fill = feature_dependent_colour
    worksheet.cell(row=9, column=1).border = border 
    
    # Set the row height for the header
    worksheet.row_dimensions[1].height = 61  # Adjust the height as needed
    
    # Set the column width for all columns
    for col in worksheet.columns:
        col_letter = col[0].column_letter
        worksheet.column_dimensions[col_letter].width = 30
    

    worksheet = workbook[enum_sheet_name]
    for col in worksheet.columns:
        col_letter = col[0].column_letter
        worksheet.column_dimensions[col_letter].width = 30
    
    # Save and close the Excel file
    workbook.save(save_path)


    return save_path


